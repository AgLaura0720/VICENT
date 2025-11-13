from pathlib import Path
from datetime import datetime
import re
import random
import time  # para tiempos de captura

import numpy as np
import pandas as pd
import serial  # 👈 nuevo: para hablar con Arduino

from openpyxl import load_workbook, Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

# ===================== CONFIG =====================

MAIN_DIR = Path(r"C:\Users\Usuario\Desktop\Uni\Cuarto Semestre\Instrumentación y Sensores\BNO055\PacienteData")
EXCEL_NAME = "Lecturas.xlsx"

# puerto y parámetros del Arduino
SERIAL_PORT = "COM9"      # 👈 CAMBIA esto si tu Arduino está en otro puerto
BAUD_RATE = 115200
DURACION_CAPTURA_S = 10   # tiempo de captura por corrida (segundos)

# ejercicio que quieres medir en ESTA ejecución:
# "fe" = Flex/Ext; "ur" = Ulnar/Radial; "ps" = Prono/Sup
EJERCICIO_ACTUAL = "fe"

COLS = [
    "timestamp_s",
    "ROM Flexión/Extensión_°",
    "EMG(F/E)_mv",
    "ROM Desviación Ulnar/Radial_°",
    "EMG(D)_mv",
    "ROM Pronosupinación_°",
    "EMG(PS)_mv",
    "Fuerza de Prensión_Kg",
    "EMG(FP)_mv"
]

EJERCICIOS = [
    ("Flexión/Extensión", "ROM Flexión/Extensión_°"),
    ("Desviación Ulnar/Radial", "ROM Desviación Ulnar/Radial_°"),
    ("Pronosupinación", "ROM Pronosupinación_°"),
    ("Fuerza de Prensión", "Fuerza de Prensión_Kg"),
]

EMG_MAP = {
    "EMG(F/E)_mv": "ROM Flexión/Extensión_°",
    "EMG(D)_mv":   "ROM Desviación Ulnar/Radial_°",
    "EMG(PS)_mv":  "ROM Pronosupinación_°",
    "EMG(FP)_mv":  "Fuerza de Prensión_Kg",
}

# ===================== FUNCIONES BASE =====================

def pedir_cedula() -> str:
    ced = input("Ingresa la cédula (solo números): ").strip()
    ced = re.sub(r"[.\s-]+", "", ced)
    if not ced.isdigit():
        raise ValueError("La cédula debe contener solo dígitos.")
    if not (7 <= len(ced) <= 12):
        raise ValueError("Longitud de cédula no válida (esperado 7–12 dígitos).")
    return ced

def ahora_nombres():
    ts = datetime.now()
    hoja = f"sesion_{ts.strftime('%Y-%m-%d_%H-%M-%S')}"[:31]
    table_name = f"TablaDatos_{ts.strftime('%H%M%S')}"
    table_name = re.sub(r'[^A-Za-z0-9_]', '_', table_name)[:31]
    return ts, hoja, table_name

def abrir_o_crear_xlsx(ruta):
    ruta.parent.mkdir(parents=True, exist_ok=True)
    if ruta.exists():
        # Intentar abrir existente
        return load_workbook(ruta)
    # Crear nuevo
    wb = Workbook()
    wb.save(ruta)
    return wb

# --------- (la función de prueba la dejo por si luego quieres simular) ----------
def generar_df_prueba(n=200) -> pd.DataFrame:
    """Datos ficticios solo para pruebas. NO se usa si capturamos de Arduino."""
    t = pd.Series(np.arange(n), dtype=float) / 50.0
    rng = random.Random(7)

    def ruido(s=1.0):
        return pd.Series([rng.uniform(-s, s) for _ in range(n)])

    rom_fe = 20*np.sin(2*np.pi*(t/6)) + 45 + ruido(1.0)
    rom_ur = 15*np.sin(2*np.pi*(t/8)) + 10 + ruido(1.0)
    rom_ps = 30*np.sin(2*np.pi*(t/7)) + ruido(1.2)
    grip   = 15 + 5*np.sin(2*np.pi*(t/5))  + ruido(0.6)

    emg_fe = (0.01*np.abs(rom_fe) + ruido(0.5)).abs()
    emg_d  = (0.012*np.abs(rom_ur) + ruido(0.4)).abs()
    emg_ps = (0.009*np.abs(rom_ps) + ruido(0.5)).abs()
    emg_fp = (0.03*np.abs(grip)    + ruido(0.6)).abs()

    df = pd.DataFrame({
        "timestamp_s": t.round(3),
        "ROM Flexión/Extensión_°": np.round(rom_fe, 2),
        "EMG(F/E)_mv": np.round(emg_fe, 3),
        "ROM Desviación Ulnar/Radial_°": np.round(rom_ur, 2),
        "EMG(D)_mv": np.round(emg_d, 3),
        "ROM Pronosupinación_°": np.round(rom_ps, 2),
        "EMG(PS)_mv": np.round(emg_ps, 3),
        "Fuerza de Prensión_Kg": np.round(grip, 2),
        "EMG(FP)_mv": np.round(emg_fp, 3),
    })
    return df
# ----------------------------------------------------------------------

def _emg_global_y_momentos(df: pd.DataFrame):
    """Máx/Mín global de todos los EMG y momento asociado (ROM/Fuerza)."""
    emg_cols = list(EMG_MAP.keys())

    max_vals = df[emg_cols].max(numeric_only=True)
    emg_max_col = max_vals.idxmax()
    emg_max_val = float(max_vals.max())
    i_max = int(df[emg_max_col].idxmax())
    assoc_col_max = EMG_MAP[emg_max_col]
    momento_max = f"{assoc_col_max} = {df.at[i_max, assoc_col_max]}"

    min_vals = df[emg_cols].min(numeric_only=True)
    emg_min_col = min_vals.idxmin()
    emg_min_val = float(min_vals.min())
    i_min = int(df[emg_min_col].idxmin())
    assoc_col_min = EMG_MAP[emg_min_col]
    momento_min = f"{assoc_col_min} = {df.at[i_min, assoc_col_min]}"

    return emg_max_val, momento_max, emg_min_val, momento_min

# ===================== LECTURA DESDE ARDUINO =====================

def _extraer_numero(linea: str):
    """
    Recibe una línea como:
      'FLEX/EXT (pitch) [deg]: -23.45'
    Devuelve: -23.45 (float) o None.
    """
    match = re.search(r"[-+]?\d*\.\d+|\d+", linea)
    if match:
        try:
            return float(match.group())
        except:
            return None
    return None


def capturar_rom_desde_arduino(puerto, ejercicio, duracion_s, baud=115200):
    """
    Habla con tu código actual de Arduino:
      - envía '1', '2' o '3' según ejercicio
      - envía ' ' (espacio) para tarar
      - lee líneas de texto, extrae el ángulo
      - construye un DataFrame con todas las columnas, dejando
        las que aún no existen en 1.

    ejercicio:
        "fe" = Flex/Ext
        "ur" = Ulnar/Radial
        "ps" = Prono/Supinación
    """
    print(f"\n📡 Abriendo puerto {puerto} a {baud} baudios...")
    ser = serial.Serial(port=puerto, baudrate=baud, timeout=1)
    time.sleep(2)  # Arduino resetea al abrir el puerto

    # Selección de ejercicio → mapeo a columna de ROM
    if ejercicio == "fe":
        cmd = "1"
        nombre_col = "ROM Flexión/Extensión_°"
    elif ejercicio == "ur":
        cmd = "2"
        nombre_col = "ROM Desviación Ulnar/Radial_°"
    elif ejercicio == "ps":
        cmd = "3"
        nombre_col = "ROM Pronosupinación_°"
    else:
        ser.close()
        raise ValueError("Ejercicio inválido: usa 'fe', 'ur' o 'ps'.")

    print(f"➡️  Seleccionando ejercicio '{ejercicio}' con comando '{cmd}'...")
    ser.write(cmd.encode())
    time.sleep(0.2)

    # Tarar con espacio (cero)
    print("➡️  Enviando TARA (espacio)...")
    ser.write(b" ")
    time.sleep(0.2)

    print(f"🎥 Capturando durante {duracion_s} segundos...")
    t0 = time.time()
    timestamps = []
    valores = []

    while (time.time() - t0) < duracion_s:
        linea = ser.readline().decode(errors="ignore").strip()
        if not linea:
            continue

        val = _extraer_numero(linea)
        if val is None:
            continue

        ts = time.time() - t0
        timestamps.append(ts)
        valores.append(val)

        print(f"[{ts:6.2f}s] {val:8.2f}")

    print("\n🛑 Enviando comando de parada ('e')...")
    ser.write(b"e")
    time.sleep(0.2)
    ser.close()

    if not valores:
        print("⚠️ No se capturó ningún dato válido.")
        return pd.DataFrame(columns=COLS)

    # Construir DataFrame: todo en 1, menos la ROM del ejercicio actual
    df = pd.DataFrame({
        "timestamp_s": timestamps,
        "ROM Flexión/Extensión_°": [1]*len(valores),
        "EMG(F/E)_mv": [1]*len(valores),
        "ROM Desviación Ulnar/Radial_°": [1]*len(valores),
        "EMG(D)_mv": [1]*len(valores),
        "ROM Pronosupinación_°": [1]*len(valores),
        "EMG(PS)_mv": [1]*len(valores),
        "Fuerza de Prensión_Kg": [1]*len(valores),
        "EMG(FP)_mv": [1]*len(valores),
    })

    df[nombre_col] = valores
    print(f"✅ Captura completada. Muestras: {len(df)}")
    return df

# ===================== GESTIÓN DE EXCEL =====================

def asegurar_inicio_simple(wb):
    """Crea hoja 'Inicio' con las dos tablas si no existe."""
    if "Inicio" not in wb.sheetnames:
        ws = wb.create_sheet("Inicio", 0)
        ws["A1"].value = "Dashboard - Resumen (simple)"
        ws["A1"].font = Font(bold=True, size=14)
        ws.append([])
        ws.append(["Fecha", "Ejercicio", "Min", "Max"])
        ws["G2"].value = "Resumen EMG global (por sesión)"
        ws["G2"].font = Font(bold=True)
        ws["G3"].value = "Fecha"
        ws["H3"].value = "Emg max"
        ws["I3"].value = "Momento del EMG max"
        ws["J3"].value = "Emg min"
        ws["K3"].value = "Momento del EMG min"
        for c in ("G", "H", "I", "J", "K"):
            ws[f"{c}3"].font = Font(bold=True)

def escribir_sesion(wb, hoja_nombre, df, table_name):
    """Crea una hoja de sesión y mete el DataFrame como tabla."""
    if hoja_nombre in wb.sheetnames:
        base = hoja_nombre
        i = 2
        while hoja_nombre in wb.sheetnames:
            hoja_nombre = (base[:28] + f"_{i}")[:31]
            i += 1

    ws = wb.create_sheet(hoja_nombre)

    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

    max_row, max_col = ws.max_row, ws.max_column
    last_col = get_column_letter(max_col)
    ref = f"A1:{last_col}{max_row}"

    tbl = Table(displayName=table_name, ref=ref)
    tbl.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
    ws.add_table(tbl)

    return hoja_nombre

def anexar_resumen_inicio(wb, ts, df):
    """Actualiza la hoja 'Inicio' con min/max por ejercicio y EMG global."""
    ws = wb["Inicio"]

    # Bloque A–D (por ejercicio)
    row = 4
    while ws.cell(row=row, column=1).value not in (None, ""):
        row += 1

    for nombre_ej, col in EJERCICIOS:
        vmin = float(pd.to_numeric(df[col], errors="coerce").min())
        vmax = float(pd.to_numeric(df[col], errors="coerce").max())
        ws.cell(row=row, column=1, value=ts.strftime("%Y-%m-%d"))
        ws.cell(row=row, column=2, value=nombre_ej)
        ws.cell(row=row, column=3, value=vmin)
        ws.cell(row=row, column=4, value=vmax)
        row += 1

    # Bloque G–K (resumen EMG global)
    row_g = 4
    while ws.cell(row=row_g, column=7).value not in (None, ""):
        row_g += 1

    emg_max_val, momento_max, emg_min_val, momento_min = _emg_global_y_momentos(df)
    ws.cell(row=row_g, column=7, value=ts.strftime("%Y-%m-%d"))
    ws.cell(row=row_g, column=8, value=emg_max_val)
    ws.cell(row=row_g, column=9, value=momento_max)
    ws.cell(row=row_g, column=10, value=emg_min_val)
    ws.cell(row=row_g, column=11, value=momento_min)

# ===================== MAIN =====================

def main():
    # 1) Pedir cédula y ruta
    paciente_id = pedir_cedula()
    ruta_xlsx = MAIN_DIR / paciente_id / EXCEL_NAME

    # 2) Capturar datos REALES desde Arduino (usa EJERCICIO_ACTUAL)
    df = capturar_rom_desde_arduino(
        puerto=SERIAL_PORT,
        ejercicio=EJERCICIO_ACTUAL,
        duracion_s=DURACION_CAPTURA_S,
        baud=BAUD_RATE
    )

    # asegurar que tenga todas las columnas y en el orden correcto
    for col in COLS:
        if col not in df.columns:
            df[col] = 1
    df = df[COLS]

    # 3) Abrir o crear Excel
    try:
        wb = abrir_o_crear_xlsx(ruta_xlsx)
    except PermissionError:
        print("❌ No se pudo abrir/crear el archivo porque está ABIERTO en Excel.")
        print("   👉 Cierra el archivo y vuelve a ejecutar el programa.")
        return

    # 4) Asegurar hoja Inicio
    asegurar_inicio_simple(wb)

    # 5) Crear hoja de sesión y actualizar Inicio
    ts, hoja, table_name = ahora_nombres()
    hoja_final = escribir_sesion(wb, hoja, df, table_name)
    anexar_resumen_inicio(wb, ts, df)

    # 6) Guardar con reintento si el Excel está abierto
    while True:
        try:
            wb.save(ruta_xlsx)
            print("\n✅ Sesión guardada correctamente.")
            print(f"   Archivo: {ruta_xlsx}")
            print(f"   Hoja de sesión: {hoja_final}")
            break
        except PermissionError:
            print("\n❌ No se pudo guardar la sesión porque el archivo Excel está ABIERTO.")
            print("   👉 Por favor, cierra el archivo 'Lecturas.xlsx' del paciente y luego pulsa Enter.")
            input("   Cuando lo hayas cerrado, presiona Enter para reintentar... ")

if __name__ == "__main__":
    main()
