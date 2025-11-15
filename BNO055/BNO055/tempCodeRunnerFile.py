from pathlib import Path
from datetime import datetime
import re
import random
import time
import numpy as np
import pandas as pd
import serial

from openpyxl import load_workbook, Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

# ===================== CONFIG =====================

MAIN_DIR = Path(r"C:\Users\14-bb0002\Desktop\PROYECTO FINAL BIOMEDICA\VICENT\BNO055\PacienteData")
EXCEL_NAME = "Lecturas.xlsx"

SERIAL_PORT = "COM4"
BAUD_RATE = 115200

COLS = [
    "timestamp_s",
    "ROM Flexión/Extensión_°",
    "EMG(F/E)_mv",
    "ROM Desviación Ulnar/Radial_°",
    "EMG(D)_mv",
    "ROM Pronosupinación_°",
    "EMG(PS)_mv",
    "Fuerza de Prensión_Kg",
    "EMG(FP)_mv"
]

EMG_MAP = {
    "EMG(F/E)_mv": "ROM Flexión/Extensión_°",
    "EMG(D)_mv":   "ROM Desviación Ulnar/Radial_°",
    "EMG(PS)_mv":  "ROM Pronosupinación_°",
    "EMG(FP)_mv":  "Fuerza de Prensión_Kg",
}

# ===================== MENÚ =====================

def menu_prueba_funcional():
    print("\n=== Selecciona la prueba funcional a realizar ===")
    print("1) P.F. Muñeca")
    print("2) P.F. Codo")
    print("3) P.F. Codo y Muñeca")
    print("----------------------------------------------")

    opcion = input("Elige una opción (1–3): ").strip()

    if opcion == "1":
        print("\n➡ Elegiste: P.F. MUÑECA")
        return {
            "nombre": "Muñeca",
            "ejercicios": [("1", "ROM Flexión/Extensión_°"),
                           ("2", "ROM Desviación Ulnar/Radial_°"),
                           ("4", "Fuerza de Prensión_Kg")]
        }

    elif opcion == "2":
        print("\n➡ Elegiste: P.F. CODO")
        return {
            "nombre": "Codo",
            "ejercicios": [("1", "ROM Flexión/Extensión_°"),
                           ("3", "ROM Pronosupinación_°"),
                           ("4", "Fuerza de Prensión_Kg")]
        }

    elif opcion == "3":
        print("\n➡ Elegiste: P.F. CODO Y MUÑECA")
        return {
            "nombre": "Codo y Muñeca",
            "ejercicios": [("1", "ROM Flexión/Extensión_°"),
                           ("2", "ROM Desviación Ulnar/Radial_°"),
                           ("3", "ROM Pronosupinación_°"),
                           ("4", "Fuerza de Prensión_Kg")]
        }

    else:
        print("❌ Opción inválida. Intenta nuevamente.")
        return menu_prueba_funcional()

# ===================== UTILIDADES =====================

def pedir_cedula() -> str:
    ced = input("Ingresa la cédula (solo números): ").strip()
    ced = re.sub(r"[.\s-]+", "", ced)
    if not ced.isdigit():
        raise ValueError("La cédula debe contener solo dígitos.")
    if not (7 <= len(ced) <= 12):
        raise ValueError("Longitud de cédula no válida.")
    return ced

def ahora_nombres():
    ts = datetime.now()
    hoja = f"sesion_{ts.strftime('%Y-%m-%d_%H-%M-%S')}"[:31]
    table_name = f"TablaDatos_{ts.strftime('%H%M%S')}"
    table_name = re.sub(r'[^A-Za-z0-9_]', '_', table_name)[:31]
    return ts, hoja, table_name

def abrir_o_crear_xlsx(ruta):
    ruta.parent.mkdir(parents=True, exist_ok=True)
    if ruta.exists():
        return load_workbook(ruta)
    wb = Workbook()
    wb.save(ruta)
    return wb

def asegurar_inicio_simple(wb):
    if "Inicio" not in wb.sheetnames:
        ws = wb.create_sheet("Inicio", 0)
        ws["A1"].value = "Dashboard - Resumen"
        ws["A1"].font = Font(bold=True, size=14)
        ws.append([])
        ws.append(["Fecha", "Ejercicio", "Min", "Max"])

def escribir_sesion(wb, hoja_nombre, df, table_name):
    if hoja_nombre in wb.sheetnames:
        base = hoja_nombre
        i = 2
        while hoja_nombre in wb.sheetnames:
            hoja_nombre = (base[:28] + f"_{i}")[:31]
            i += 1

    ws = wb.create_sheet(hoja_nombre)

    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

    max_row, max_col = ws.max_row, ws.max_column
    last_col = get_column_letter(max_col)
    ref = f"A1:{last_col}{max_row}"

    tbl = Table(displayName=table_name, ref=ref)
    tbl.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
    ws.add_table(tbl)

    return hoja_nombre

# ===================== CAPTURA ARDUINO =====================

def _extraer_numero(linea: str):
    match = re.search(r"[-+]?\d*\.\d+|\d+", linea)
    return float(match.group()) if match else None

def capturar_rom_desde_arduino(cmd, nombre_col):
    duracion = int(input(f"Tiempo de captura para {nombre_col}: "))

    print(f"\n📡 Abriendo puerto {SERIAL_PORT}...")
    ser = serial.Serial(port=SERIAL_PORT, baudrate=BAUD_RATE, timeout=1)
    time.sleep(2)

    print(f"➡️ Enviando comando '{cmd}'...")
    ser.write(cmd.encode())
    time.sleep(0.3)

    print("➡️ TARA enviada...")
    ser.write(b" ")
    time.sleep(0.3)

    t0 = time.time()
    timestamps, valores = [], []

    print(f"🎥 Capturando {duracion} segundos...\n")

    while (time.time() - t0) < duracion:
        linea = ser.readline().decode(errors="ignore").strip()
        val = _extraer_numero(linea)
        if val is None:
            continue
        ts = time.time() - t0
        timestamps.append(ts)
        valores.append(val)
        print(f"[{ts:6.2f}s] {val:8.2f}")

    print("\n🛑 Enviando 'e'...")
    ser.write(b"e")
    ser.close()

    df = pd.DataFrame({c: [1]*len(valores) for c in COLS})
    df["timestamp_s"] = timestamps
    df[nombre_col] = valores
    return df

# ===================== MAIN =====================

def main():
    pf = menu_prueba_funcional()
    paciente_id = pedir_cedula()

    ruta_xlsx = MAIN_DIR / paciente_id / EXCEL_NAME

    try:
        wb = abrir_o_crear_xlsx(ruta_xlsx)
    except PermissionError:
        print("❌ Cierra el Excel e intenta de nuevo.")
        return

    asegurar_inicio_simple(wb)

    ts, hoja, table_name = ahora_nombres()
    lista_dfs = []

    for cmd, nombre_col in pf["ejercicios"]:
        print(f"\n=== Capturando: {nombre_col} ===")
        df_ej = capturar_rom_desde_arduino(cmd, nombre_col)
        lista_dfs.append(df_ej)

    # Unir todas las capturas de la sesión
    df_final = pd.concat(lista_dfs, ignore_index=True)

    # Escribir sesión en Excel
    hoja_final = escribir_sesion(wb, hoja, df_final, table_name)

    # Guardar archivo
    wb.save(ruta_xlsx)

    print("\n✅ Sesión guardada correctamente.")
    print(f"Archivo: {ruta_xlsx}")
    print(f"Hoja creada: {hoja_final}")


if __name__ == "__main__":
    main()
