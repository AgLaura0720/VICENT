# This Python file uses the following encoding: utf-8
# scripts/arduino_controller.py
# Wrapper non-interactive para tu código original.
# Lee comandos por stdin y controla la captura (serial) y guardado (.xlsx).
# Copia en scripts/arduino_controller.py

from pathlib import Path
from datetime import datetime
import re
import time
import sys
import threading

import numpy as np
import pandas as pd
import serial

from openpyxl import load_workbook, Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

# ---------------- CONFIG (ajusta si hace falta) ----------------
MAIN_DIR = Path(r"C:\Users\Adrian Jr\Desktop\VICENT\BNO055\PacienteData")
EXCEL_NAME = "Lecturas.xlsx"

SERIAL_PORT = "COM4"
BAUD_RATE = 115200
SERIAL_TIMEOUT = 1.0

COLS = [
    "timestamp_s",
    "ROM Flexión/Extensión_°",
    "EMG(F/E)_mv",
    "ROM Desviación Ulnar/Radial_°",
    "EMG(D)_mv",
    "ROM Pronosupinación_°",
    "EMG(PS)_mv",
    "Fuerza de Prensión_Kg",
    "EMG(FP)_mv"
]

EMG_MAP = {
    "EMG(F/E)_mv": "ROM Flexión/Extensión_°",
    "EMG(D)_mv":   "ROM Desviación Ulnar/Radial_°",
    "EMG(PS)_mv":  "ROM Pronosupinación_°",
    "EMG(FP)_mv":  "Fuerza de Prensión_Kg",
}

# ---------------- utilidades (copiadas/adaptadas) ----------------

def ahora_nombres():
    ts = datetime.now()
    hoja = f"sesion_{ts.strftime('%Y-%m-%d_%H-%M-%S')}"[:31]
    table_name = f"TablaDatos_{ts.strftime('%H%M%S')}"
    table_name = re.sub(r'[^A-Za-z0-9_]', '_', table_name)[:31]
    return ts, hoja, table_name

def abrir_o_crear_xlsx(ruta):
    ruta.parent.mkdir(parents=True, exist_ok=True)
    if ruta.exists():
        return load_workbook(ruta)
    wb = Workbook()
    wb.save(ruta)
    return wb

def asegurar_inicio_simple(wb):
    if "Inicio" not in wb.sheetnames:
        ws = wb.create_sheet("Inicio", 0)
        ws["A1"].value = "Dashboard - Resumen"
        ws["A1"].font = Font(bold=True, size=14)
        ws.append([])
        ws.append(["Fecha", "Ejercicio", "Min", "Max"])

def escribir_sesion(wb, hoja_nombre, df, table_name):
    if hoja_nombre in wb.sheetnames:
        base = hoja_nombre
        i = 2
        while hoja_nombre in wb.sheetnames:
            hoja_nombre = (base[:28] + f"_{i}")[:31]
            i += 1

    ws = wb.create_sheet(hoja_nombre)

    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

    max_row, max_col = ws.max_row, ws.max_column
    last_col = get_column_letter(max_col)
    ref = f"A1:{last_col}{max_row}"

    tbl = Table(displayName=table_name, ref=ref)
    tbl.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
    ws.add_table(tbl)

    return hoja_nombre

# ---------------- capturar desde Arduino (adaptada) ----------------

def _extraer_numero(linea: str):
    match = re.search(r"[-+]?\d*\.\d+|\d+", linea)
    return float(match.group()) if match else None

def capturar_rom_desde_arduino(cmd: str, nombre_col: str, duracion: int, serial_port=SERIAL_PORT, baud=BAUD_RATE):
    """
    Ejecuta una captura no interactiva:
      cmd: comando que se enviará por Serial (ej "1")
      nombre_col: nombre de columna (ej "ROM Flexión/Extensión_°")
      duracion: segundos de captura
    Devuelve DataFrame con columnas (COLS) y la columna `nombre_col` llena.
    Mientras captura imprime por stdout líneas máquina-amigables:
      DATA:<colname>,<timestamp_s>,<value>
    """
    try:
        ser = serial.Serial(port=serial_port, baudrate=baud, timeout=SERIAL_TIMEOUT)
    except Exception as e:
        print(f"ERROR:SERIAL_OPEN:{e}", flush=True)
        return None

    # dar tiempo a Arduino
    time.sleep(0.2)

    # enviar comando al Arduino (ej "1" o "1:5" si preferimos)
    ser.write((cmd + "\n").encode())
    ser.flush()
    time.sleep(0.05)

    # enviar TARA (si tu firmware lo usa)
    ser.write(b" ")
    ser.flush()
    time.sleep(0.05)

    t0 = time.time()
    timestamps, valores = [], []

    print(f"STATUS:CAPTURE_STARTED:{nombre_col}", flush=True)

    while (time.time() - t0) < duracion:
        try:
            linea = ser.readline().decode(errors="ignore").strip()
        except Exception:
            linea = ""
        if not linea:
            continue

        # Intenta extraer número como antes
        val = _extraer_numero(linea)
        if val is None:
            # si la línea contiene mensajes del Arduino podemos reenviarlos por stdout
            # por ejemplo: CAPTURE_START, END, etc
            print(f"HWMSG:{linea}", flush=True)
            continue

        ts = time.time() - t0
        timestamps.append(ts)
        valores.append(val)

        # Emitir línea máquina-amigable para que la UI muestre en tiempo real
        print(f"DATA:{nombre_col},{ts:.3f},{val:.6f}", flush=True)

    # señal de fin al Arduino (como tu Python hacía)
    try:
        ser.write(b"e")
        ser.flush()
    except Exception:
        pass
    ser.close()

    # crear DataFrame con la estructura de COLS
    df = pd.DataFrame({c: [1]*len(valores) for c in COLS})
    df["timestamp_s"] = timestamps
    df[nombre_col] = valores
    print(f"STATUS:CAPTURE_END:{nombre_col}", flush=True)
    return df

# ---------------- controlador por stdin ----------------

class Controller:
    def __init__(self):
        self.patient_id = None
        self.session_dfs = []  # lista de dataframes por ejercicio en la sesión
        self.serial_port = SERIAL_PORT
        self.baud = BAUD_RATE
        print("STATUS:READY", flush=True)

    def handle_line(self, line: str):
        line = line.strip()
        if not line:
            return

        # Comandos simples
        if line.upper().startswith("PATIENT:"):
            _, val = line.split(":", 1)
            val = re.sub(r"[.\s-]+", "", val)
            self.patient_id = val
            print(f"STATUS:PATIENT_SET:{self.patient_id}", flush=True)
            return

        if line.upper().startswith("START:"):
            # formato START:cmd:colname:dur
            parts = line.split(":", 3)
            if len(parts) < 4:
                print("ERROR:START_FORMAT", flush=True)
                return
            _, cmd, colname, dur_s = parts
            try:
                dur = int(dur_s)
            except:
                print("ERROR:DURATION", flush=True)
                return
            # iniciar captura (bloqueante) y añadir DF a session_dfs
            df = capturar_rom_desde_arduino(cmd, colname, dur, serial_port=self.serial_port, baud=self.baud)
            if df is not None:
                self.session_dfs.append(df)
            return

        if line.upper() == "SAVE":
            # guarda todas las sesiones acumuladas en un archivo xlsx en MAIN_DIR/patient_id/
            if not self.patient_id:
                print("ERROR:NO_PATIENT", flush=True)
                return
            ruta_xlsx = MAIN_DIR / self.patient_id / EXCEL_NAME
            try:
                wb = abrir_o_crear_xlsx(ruta_xlsx)
            except PermissionError:
                print("ERROR:EXCEL_LOCKED", flush=True)
                return
            asegurar_inicio_simple(wb)
            ts, hoja, table_name = ahora_nombres()
            if not self.session_dfs:
                print("ERROR:NO_DATA", flush=True)
                return
            df_final = pd.concat(self.session_dfs, ignore_index=True)
            hoja_final = escribir_sesion(wb, hoja, df_final, table_name)
            try:
                wb.save(ruta_xlsx)
                print(f"SAVED:{ruta_xlsx}", flush=True)
                # limpiar lista de dfs después de guardar
                self.session_dfs = []
            except Exception as e:
                print(f"ERROR:SAVE_FAILED:{e}", flush=True)
            return

        if line.upper() == "STATUS":
            print("STATUS:READY", flush=True)
            return

        if line.upper() == "EXIT":
            print("STATUS:EXITING", flush=True)
            sys.exit(0)

        print(f"ERROR:UNKNOWN_CMD:{line}", flush=True)


def stdin_reader(controller: Controller):
    # Lee stdin línea a línea y la pasa al controller
    while True:
        raw = sys.stdin.readline()
        if raw == "":
            # EOF -> salir
            break
        controller.handle_line(raw)

def main():
    ctrl = Controller()
    # Ejecutamos el reader en el mismo hilo (bloqueante) para simplicidad
    try:
        stdin_reader(ctrl)
    except Exception as e:
        print(f"ERROR:CRASH:{e}", flush=True)
        sys.exit(1)

if __name__ == "__main__":
    main()

# if __name__ == "__main__":
#     pass
