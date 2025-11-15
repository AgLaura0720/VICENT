# scripts/principal.py
"""
Versión dual de principal.py:
- Si se ejecuta normalmente: corre en modo interactivo (igual que tu script original).
- Si se ejecuta y recibe comandos por stdin (por ejemplo desde QProcess en Qt),
  actúa como bridge y acepta comandos:
    PATIENT:<cedula>
    START:<cmd>:<colname>:<dur_seconds>
    SAVE
    STATUS
    EXIT

Emite por stdout mensajes máquina-amigables:
  STATUS:READY
  STATUS:CAPTURE_STARTED:<colname>
  DATA:<colname>,<timestamp_s>,<value>
  STATUS:CAPTURE_END:<colname>
  SAVED:<ruta>
  ERROR:...
  HWMSG:...
"""

from pathlib import Path
from datetime import datetime
import re
import time
import sys
import threading

import numpy as np
import pandas as pd
import serial

from openpyxl import load_workbook, Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

# ===================== CONFIG =====================

MAIN_DIR = Path(r"C:\Users\Adrian Jr\Desktop\VICENT\BNO055\PacienteData")
EXCEL_NAME = "Lecturas.xlsx"

SERIAL_PORT = "COM4"
BAUD_RATE = 115200
SERIAL_TIMEOUT = 1.0

COLS = [
    "timestamp_s",
    "ROM Flexión/Extensión_°",
    "EMG(F/E)_mv",
    "ROM Desviación Ulnar/Radial_°",
    "EMG(D)_mv",
    "ROM Pronosupinación_°",
    "EMG(PS)_mv",
    "Fuerza de Prensión_Kg",
    "EMG(FP)_mv"
]

EJERCICIOS = [
    ("Flexión/Extensión", "ROM Flexión/Extensión_°"),
    ("Desviación Ulnar/Radial", "ROM Desviación Ulnar/Radial_°"),
    ("Pronosupinación", "ROM Pronosupinación_°"),
    ("Fuerza de Prensión", "Fuerza de Prensión_Kg"),
]

EMG_MAP = {
    "EMG(F/E)_mv": "ROM Flexión/Extensión_°",
    "EMG(D)_mv":   "ROM Desviación Ulnar/Radial_°",
    "EMG(PS)_mv":  "ROM Pronosupinación_°",
    "EMG(FP)_mv":  "Fuerza de Prensión_Kg",
}

# ===================== UTILIDADES =====================

def ahora_nombres():
    ts = datetime.now()
    hoja = f"sesion_{ts.strftime('%Y-%m-%d_%H-%M-%S')}"[:31]
    table_name = f"TablaDatos_{ts.strftime('%H%M%S')}"
    table_name = re.sub(r'[^A-Za-z0-9_]', '_', table_name)[:31]
    return ts, hoja, table_name

def abrir_o_crear_xlsx(ruta):
    ruta.parent.mkdir(parents=True, exist_ok=True)
    if ruta.exists():
        return load_workbook(ruta)
    wb = Workbook()
    wb.save(ruta)
    return wb

def asegurar_inicio_simple(wb):
    if "Inicio" not in wb.sheetnames:
        ws = wb.create_sheet("Inicio", 0)
        ws["A1"].value = "Dashboard - Resumen (simple)"
        ws["A1"].font = Font(bold=True, size=14)
        ws.append([])
        ws.append(["Fecha", "Ejercicio", "Min", "Max"])
        ws["G2"].value = "Resumen EMG global (por sesión)"
        ws["G2"].font = Font(bold=True)
        ws["G3"].value = "Fecha"
        ws["H3"].value = "Emg max"
        ws["I3"].value = "Momento del EMG max"
        ws["J3"].value = "Emg min"
        ws["K3"].value = "Momento del EMG min"
        for c in ("G", "H", "I", "J", "K"):
            ws[f"{c}3"].font = Font(bold=True)

def _emg_global_y_momentos(df: pd.DataFrame):
    emg_cols = list(EMG_MAP.keys())
    emg_numeric = df[emg_cols].apply(pd.to_numeric, errors="coerce")
    valid_cols = [c for c in emg_cols if emg_numeric[c].notna().any()]

    if not valid_cols:
        return None, None, None, None

    max_vals = emg_numeric[valid_cols].max()
    emg_max_col = max_vals.idxmax()
    emg_max_val = float(max_vals.max())
    i_max = int(emg_numeric[emg_max_col].idxmax())
    assoc_col_max = EMG_MAP[emg_max_col]
    momento_max = f"{assoc_col_max} = {df.at[i_max, assoc_col_max]}"

    min_vals = emg_numeric[valid_cols].min()
    emg_min_col = min_vals.idxmin()
    emg_min_val = float(min_vals.min())
    i_min = int(emg_numeric[emg_min_col].idxmin())
    assoc_col_min = EMG_MAP[emg_min_col]
    momento_min = f"{assoc_col_min} = {df.at[i_min, assoc_col_min]}"

    return emg_max_val, momento_max, emg_min_val, momento_min

def anexar_resumen_inicio(wb, ts, df):
    ws = wb["Inicio"]
    row = 4
    while ws.cell(row=row, column=1).value not in (None, ""):
        row += 1

    for nombre_ej, col in EJERCICIOS:
        if col not in df.columns:
            continue
        serie = pd.to_numeric(df[col], errors="coerce")
        if serie.dropna().empty:
            continue
        vmin = float(serie.min())
        vmax = float(serie.max())
        ws.cell(row=row, column=1, value=ts.strftime("%Y-%m-%d"))
        ws.cell(row=row, column=2, value=nombre_ej)
        ws.cell(row=row, column=3, value=vmin)
        ws.cell(row=row, column=4, value=vmax)
        row += 1

    row_g = 4
    while ws.cell(row=row_g, column=7).value not in (None, ""):
        row_g += 1

    emg_max_val, momento_max, emg_min_val, momento_min = _emg_global_y_momentos(df)
    if emg_max_val is None:
        return
    ws.cell(row=row_g, column=7, value=ts.strftime("%Y-%m-%d"))
    ws.cell(row=row_g, column=8, value=emg_max_val)
    ws.cell(row=row_g, column=9, value=momento_max)
    ws.cell(row=row_g, column=10, value=emg_min_val)
    ws.cell(row=row_g, column=11, value=momento_min)

def escribir_sesion(wb, hoja_nombre, df, table_name):
    if hoja_nombre in wb.sheetnames:
        base = hoja_nombre
        i = 2
        while hoja_nombre in wb.sheetnames:
            hoja_nombre = (base[:28] + f"_{i}")[:31]
            i += 1

    ws = wb.create_sheet(hoja_nombre)

    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

    max_row, max_col = ws.max_row, ws.max_column
    last_col = get_column_letter(max_col)
    ref = f"A1:{last_col}{max_row}"

    tbl = Table(displayName=table_name, ref=ref)
    tbl.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
    ws.add_table(tbl)

    return hoja_nombre

# ===================== SERIAL / ARDUINO CAPTURE =====================

def _extraer_numero(linea: str):
    match = re.search(r"[-+]?\d*\.\d+|\d+", linea)
    return float(match.group()) if match else None

def capturar_rom_desde_arduino(cmd, nombre_col, duracion=None, serial_port=SERIAL_PORT, baud=BAUD_RATE):
    """
    Si duracion is None, se pedirá por input() (modo interactivo).
    Si duracion es int, se usará (modo bridge).
    """
    if duracion is None:
        duracion = int(input(f"Tiempo de captura para {nombre_col}: "))

    try:
        ser = serial.Serial(port=serial_port, baudrate=baud, timeout=1)
    except Exception as e:
        print(f"ERROR:SERIAL_OPEN:{e}", flush=True)
        return None

    time.sleep(0.2)

    # Enviar comando simple (arduino acepta "1" o "1:5" etc; aquí enviamos solo el número)
    ser.write(cmd.encode())
    ser.flush()
    time.sleep(0.05)

    # Enviar TARA como haces en el sketch
    ser.write(b" ")
    ser.flush()
    time.sleep(0.05)

    t0 = time.time()
    timestamps, valores = [], []

    print(f"STATUS:CAPTURE_STARTED:{nombre_col}", flush=True)

    while (time.time() - t0) < duracion:
        try:
            linea = ser.readline().decode(errors="ignore").strip()
        except Exception:
            linea = ""
        if not linea:
            continue

        # Reenvía mensajes no numéricos como HWMSG
        val = _extraer_numero(linea)
        if val is None:
            print(f"HWMSG:{linea}", flush=True)
            continue

        ts = time.time() - t0
        timestamps.append(ts)
        valores.append(val)

        # Línea máquina-amigable
        print(f"DATA:{nombre_col},{ts:.3f},{val:.6f}", flush=True)

    # notify end, enviar 'e' al Arduino
    try:
        ser.write(b"e")
        ser.flush()
    except Exception:
        pass
    ser.close()

    print(f"STATUS:CAPTURE_END:{nombre_col}", flush=True)

    # Construir DF consistente con tu pipeline
    df = pd.DataFrame({
        "timestamp_s": timestamps,
        nombre_col: valores
    })

    return df

# ===================== CONTROLLER (stdin bridge) =====================

class Controller:
    def __init__(self):
        self.patient_id = None
        self.session_dfs = []
        self.serial_port = SERIAL_PORT
        self.baud = BAUD_RATE
        print("STATUS:READY", flush=True)

    def handle_line(self, line: str):
        line = line.strip()
        if not line:
            return

        if line.upper().startswith("PATIENT:"):
            _, val = line.split(":", 1)
            val = re.sub(r"[.\s-]+", "", val)
            self.patient_id = val
            print(f"STATUS:PATIENT_SET:{self.patient_id}", flush=True)
            return

        if line.upper().startswith("START:"):
            parts = line.split(":", 3)
            if len(parts) < 4:
                print("ERROR:START_FORMAT", flush=True)
                return
            _, cmd, colname, dur_s = parts
            try:
                dur = int(dur_s)
            except:
                print("ERROR:DURATION", flush=True)
                return
            df = capturar_rom_desde_arduino(cmd, colname, dur, serial_port=self.serial_port, baud=self.baud)
            if df is not None:
                self.session_dfs.append(df)
            return

        if line.upper() == "SAVE":
            if not self.patient_id:
                print("ERROR:NO_PATIENT", flush=True)
                return
            ruta_xlsx = MAIN_DIR / self.patient_id / EXCEL_NAME
            try:
                wb = abrir_o_crear_xlsx(ruta_xlsx)
            except PermissionError:
                print("ERROR:EXCEL_LOCKED", flush=True)
                return
            asegurar_inicio_simple(wb)
            ts, hoja, table_name = ahora_nombres()
            if not self.session_dfs:
                print("ERROR:NO_DATA", flush=True)
                return
            # combinar (como en tu script original)
            df_final = self._combine_session_dfs()
            hoja_final = escribir_sesion(wb, hoja, df_final, table_name)
            anexar_resumen_inicio(wb, ts, df_final)
            try:
                wb.save(ruta_xlsx)
                print(f"SAVED:{ruta_xlsx}", flush=True)
                self.session_dfs = []
            except Exception as e:
                print(f"ERROR:SAVE_FAILED:{e}", flush=True)
            return

        if line.upper() == "STATUS":
            print("STATUS:READY", flush=True)
            return

        if line.upper() == "EXIT":
            print("STATUS:EXITING", flush=True)
            sys.exit(0)

        print(f"ERROR:UNKNOWN_CMD:{line}", flush=True)

    def _combine_session_dfs(self):
        # Reproduce la lógica de combinación de tu main original
        if not self.session_dfs:
            return pd.DataFrame(columns=COLS)

        df_final = self.session_dfs[0].copy()
        # ensure all others are merged as in your script
        for df_ej in self.session_dfs[1:]:
            max_len = max(len(df_final), len(df_ej))
            df_final = df_final.reindex(range(max_len))
            df_ej    = df_ej.reindex(range(max_len))
            # find the column in df_ej other than timestamp and merge
            for c in df_ej.columns:
                if c == "timestamp_s": continue
                df_final[c] = df_ej[c]
        # ensure all COLS present
        for c in COLS:
            if c not in df_final.columns:
                df_final[c] = np.nan
        df_final = df_final[COLS]
        return df_final

def stdin_reader(controller: Controller):
    while True:
        raw = sys.stdin.readline()
        if raw == "":
            break
        controller.handle_line(raw)

# ===================== MODO INTERACTIVO (tu main original) =====================

def menu_prueba_funcional():
    print("\n=== Selecciona la prueba funcional a realizar ===")
    print("1) P.F. Muñeca")
    print("2) P.F. Codo")
    print("3) P.F. Codo y Muñeca")
    print("----------------------------------------------")
    opcion = input("Elige una opción (1–3): ").strip()
    if opcion == "1":
        print("\n➡ Elegiste: P.F. MUÑECA")
        return {
            "nombre": "Muñeca",
            "ejercicios": [("1", "ROM Flexión/Extensión_°"),
                           ("2", "ROM Desviación Ulnar/Radial_°"),
                           ("4", "Fuerza de Prensión_Kg")]
        }
    elif opcion == "2":
        print("\n➡ Elegiste: P.F. CODO")
        return {
            "nombre": "Codo",
            "ejercicios": [("1", "ROM Flexión/Extensión_°"),
                           ("3", "ROM Pronosupinación_°"),
                           ("4", "Fuerza de Prensión_Kg")]
        }
    elif opcion == "3":
        print("\n➡ Elegiste: P.F. CODO Y MUÑECA")
        return {
            "nombre": "Codo y Muñeca",
            "ejercicios": [("1", "ROM Flexión/Extensión_°"),
                           ("2", "ROM Desviación Ulnar/Radial_°"),
                           ("3", "ROM Pronosupinación_°"),
                           ("4", "Fuerza de Prensión_Kg")]
        }
    else:
        print("❌ Opción inválida. Intenta nuevamente.")
        return menu_prueba_funcional()

def pedir_cedula_interactiva() -> str:
    ced = input("Ingresa la cédula (solo números): ").strip()
    ced = re.sub(r"[.\s-]+", "", ced)
    if not ced.isdigit():
        raise ValueError("La cédula debe contener solo dígitos.")
    if not (7 <= len(ced) <= 12):
        raise ValueError("Longitud de cédula no válida.")
    return ced

def main_interactive():
    pf = menu_prueba_funcional()
    paciente_id = pedir_cedula_interactiva()
    ruta_xlsx = MAIN_DIR / paciente_id / EXCEL_NAME
    try:
        wb = abrir_o_crear_xlsx(ruta_xlsx)
    except PermissionError:
        print("❌ Cierra el Excel e intenta de nuevo.")
        return
    asegurar_inicio_simple(wb)
    ts, hoja , table_name = ahora_nombres()
    lista_dfs = []
    for cmd, nombre_col in pf["ejercicios"]:
        print(f"\n=== Capturando: {nombre_col} ===")
        df_ej = capturar_rom_desde_arduino(cmd, nombre_col)
        lista_dfs.append(df_ej)
    # combinar
    df_final = lista_dfs[0].copy()
    for df_ej, (cmd, nombre_col) in zip(lista_dfs[1:], pf["ejercicios"][1:]):
        max_len = max(len(df_final), len(df_ej))
        df_final = df_final.reindex(range(max_len))
        df_ej    = df_ej.reindex(range(max_len))
        df_final[nombre_col] = df_ej[nombre_col]
    for c in COLS:
        if c not in df_final.columns:
            df_final[c] = np.nan
    df_final = df_final[COLS]
    hoja_final = escribir_sesion(wb, hoja, df_final, table_name)
    anexar_resumen_inicio(wb, ts, df_final)
    wb.save(ruta_xlsx)
    print("\n✅ Sesión guardada correctamente.")
    print(f"Archivo: {ruta_xlsx}")
    print(f"Hoja creada: {hoja_final}")

# ===================== ENTRY POINT =====================

def main_bridge_mode():
    ctrl = Controller()
    try:
        stdin_reader(ctrl)
    except Exception as e:
        print(f"ERROR:CRASH:{e}", flush=True)
        sys.exit(1)

if __name__ == "__main__":
    # Detectar si stdin no es tty -> probablemente bridge (QProcess)
    # O si pasas un argumento --bridge, forzamos modo bridge.
    is_bridge = (not sys.stdin.isatty()) or ("--bridge" in sys.argv)
    if is_bridge:
        main_bridge_mode()
    else:
        main_interactive()
