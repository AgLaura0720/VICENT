#!/usr/bin/env python
from pathlib import Path
from datetime import datetime
import sys
import re

import numpy as np
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

# Columnas finales, igual que tu script original
COLS = [
    "timestamp_s",
    "ROM Flexión/Extensión_°",
    "EMG(F/E)_mv",
    "ROM Desviación Ulnar/Radial_°",
    "EMG(D)_mv",
    "ROM Pronosupinación_°",
    "EMG(PS)_mv",
    "Fuerza de Prensión_Kg",
    "EMG(FP)_mv",
]

EJERCICIOS = [
    ("Flexión / Extensión", "ROM Flexión/Extensión_°"),
    ("Desviación cubital/radial", "ROM Desviación Ulnar/Radial_°"),
    ("Pronosupinación", "ROM Pronosupinación_°"),
    ("Fuerza de prensión", "Fuerza de Prensión_Kg"),
]

EMG_MAP = {
    "EMG(F/E)_mv": "ROM Flexión/Extensión_°",
    "EMG(D)_mv":   "ROM Desviación Ulnar/Radial_°",
    "EMG(PS)_mv":  "ROM Pronosupinación_°",
    "EMG(FP)_mv":  "Fuerza de Prensión_Kg",
}

def ahora_nombres():
    ts = datetime.now()
    hoja = f"sesion_{ts.strftime('%Y-%m-%d_%H-%M-%S')}"[:31]
    table_name = f"TablaDatos_{ts.strftime('%H%M%S')}"
    table_name = re.sub(r"[^A-Za-z0-9_]", "_", table_name)[:31]
    return ts, hoja, table_name

def abrir_o_crear_xlsx(ruta: Path):
    ruta.parent.mkdir(parents=True, exist_ok=True)
    if ruta.exists():
        return load_workbook(ruta)
    wb = Workbook()
    wb.save(ruta)
    return wb

def asegurar_inicio_simple(wb):
    if "Inicio" not in wb.sheetnames:
        ws = wb.create_sheet("Inicio", 0)
        ws["A1"].value = "Dashboard - Resumen (simple)"
        ws["A1"].font = Font(bold=True, size=14)
        ws.append([])
        ws.append(["Fecha", "Ejercicio", "Min", "Max"])
        ws["G2"].value = "Resumen EMG global (por sesión)"
        ws["G2"].font = Font(bold=True)
        ws["G3"].value = "Fecha"
        ws["H3"].value = "Emg max"
        ws["I3"].value = "Momento del EMG max"
        ws["J3"].value = "Emg min"
        ws["K3"].value = "Momento del EMG min"
        for c in ("G", "H", "I", "J", "K"):
            ws[f"{c}3"].font = Font(bold=True)

def _emg_global_y_momentos(df: pd.DataFrame):
    emg_cols = list(EMG_MAP.keys())
    emg_numeric = df[emg_cols].apply(pd.to_numeric, errors="coerce")
    valid_cols = [c for c in emg_cols if emg_numeric[c].notna().any()]
    if not valid_cols:
        return None, None, None, None

    max_vals = emg_numeric[valid_cols].max()
    emg_max_col = max_vals.idxmax()
    emg_max_val = float(max_vals.max())
    i_max = int(emg_numeric[emg_max_col].idxmax())
    assoc_col_max = EMG_MAP[emg_max_col]
    momento_max = f"{assoc_col_max} = {df.at[i_max, assoc_col_max]}"

    min_vals = emg_numeric[valid_cols].min()
    emg_min_col = min_vals.idxmin()
    emg_min_val = float(min_vals.min())
    i_min = int(emg_numeric[emg_min_col].idxmin())
    assoc_col_min = EMG_MAP[emg_min_col]
    momento_min = f"{assoc_col_min} = {df.at[i_min, assoc_col_min]}"

    return emg_max_val, momento_max, emg_min_val, momento_min

def anexar_resumen_inicio(wb, ts: datetime, df: pd.DataFrame):
    ws = wb["Inicio"]

    # Bloque A–D
    row = 4
    while ws.cell(row=row, column=1).value not in (None, ""):
        row += 1

    for nombre_ej, col in EJERCICIOS:
        if col not in df.columns:
            continue
        serie = pd.to_numeric(df[col], errors="coerce")
        if serie.dropna().empty:
            continue
        vmin = float(serie.min())
        vmax = float(serie.max())
        ws.cell(row=row, column=1, value=ts.strftime("%Y-%m-%d"))
        ws.cell(row=row, column=2, value=nombre_ej)
        ws.cell(row=row, column=3, value=vmin)
        ws.cell(row=row, column=4, value=vmax)
        row += 1

    # Bloque G–K (EMG global)
    row_g = 4
    while ws.cell(row=row_g, column=7).value not in (None, ""):
        row_g += 1

    emg_max_val, momento_max, emg_min_val, momento_min = _emg_global_y_momentos(df)
    if emg_max_val is None:
        return

    ws.cell(row=row_g, column=7, value=ts.strftime("%Y-%m-%d"))
    ws.cell(row=row_g, column=8, value=emg_max_val)
    ws.cell(row=row_g, column=9, value=momento_max)
    ws.cell(row=row_g, column=10, value=emg_min_val)
    ws.cell(row=row_g, column=11, value=momento_min)

def escribir_sesion(wb, hoja_nombre: str, df: pd.DataFrame, table_name: str):
    if hoja_nombre in wb.sheetnames:
        base = hoja_nombre
        i = 2
        while hoja_nombre in wb.sheetnames:
            hoja_nombre = (base[:28] + f"_{i}")[:31]
            i += 1
    ws = wb.create_sheet(hoja_nombre)
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)
    max_row, max_col = ws.max_row, ws.max_column
    last_col = get_column_letter(max_col)
    ref = f"A1:{last_col}{max_row}"
    tbl = Table(displayName=table_name, ref=ref)
    tbl.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
    ws.add_table(tbl)
    return hoja_nombre

def main(argv):
    if len(argv) < 4:
        print("Uso: build_excel_from_csvs.py <cedula> <nombre_examen> <csv1> [csv2 csv3 ...]")
        return 1

    patient_id = argv[1]
    exam_name  = argv[2]
    csv_paths  = [Path(p) for p in argv[3:]]

    # Leemos cada CSV (timestamp_s + una columna ROM)
    lista_dfs = []
    for p in csv_paths:
        df = pd.read_csv(p)
        lista_dfs.append(df)

    if not lista_dfs:
        print("No hay CSV para procesar.")
        return 1

    # Unir como en tu script original
    df_final = lista_dfs[0].copy()
    for df_ej in lista_dfs[1:]:
        max_len = max(len(df_final), len(df_ej))
        df_final = df_final.reindex(range(max_len))
        df_ej    = df_ej.reindex(range(max_len))

        # Copiar todas las columnas excepto timestamp_s
        for col in df_ej.columns:
            if col == "timestamp_s":
                continue
            df_final[col] = df_ej[col].values

    # Asegurar columnas finales
    for c in COLS:
        if c not in df_final.columns:
            df_final[c] = np.nan
    df_final = df_final[COLS]


    # Ruta de Lecturas.xlsx = carpeta del primer CSV / Lecturas.xlsx
    base_dir = csv_paths[0].parent
    xlsx_path = base_dir / "Lecturas.xlsx"

    wb = abrir_o_crear_xlsx(xlsx_path)
    asegurar_inicio_simple(wb)

    ts, hoja_nombre, table_name = ahora_nombres()
    escribir_sesion(wb, hoja_nombre, df_final, table_name)
    anexar_resumen_inicio(wb, ts, df_final)

    wb.save(xlsx_path)
    print(f"Guardado Excel en {xlsx_path}")
    return 0

if __name__ == "__main__":
    raise SystemExit(main(sys.argv))
